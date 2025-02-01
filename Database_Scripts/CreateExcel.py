from Models import DatabaseModels

import openpyxl

class CreateExcelFile:

    def __init__(self):
        self.database = DatabaseModels()
        self.wbPath = "C:/Users/mulim/OneDrive/Área de Trabalho/Projeto Spotify/ExportMyTimeMachine.xlsx"

        self.createWorkbook()

    def createWorkbook(self):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        workbook.remove(sheet)

        workbook.create_sheet("Musics")
        workbook["Musics"]["A1"].value = "ID"
        workbook["Musics"]["B1"].value = "Title"
        workbook["Musics"]["C1"].value = "Artists"
        workbook["Musics"]["D1"].value = "Duration"
        workbook["Musics"]["E1"].value = "Album_Img_Url"

        workbook.create_sheet("Executions")
        workbook["Executions"]["A1"].value = "ID"
        workbook["Executions"]["B1"].value = "Music_ID"
        workbook["Executions"]["C1"].value = "Played_At"

        workbook.save(self.wbPath)

    def execute(self):
        
        wb = openpyxl.load_workbook(self.wbPath)
        musicSheet = wb["Musics"]
        ExecutionSheet = wb["Executions"]

        musicLines = self.database.readRows("Music")
        ExecutionLines = self.database.readRows("Executions")

        row = 2
        for offset in range(0, musicLines, 100):
            response = self.database.readAllRows("Music", limit=100, offset=offset)

            for music in response:

                musicSheet[f"A{row}"].value = music.id
                musicSheet[f"B{row}"].value = music.title
                musicSheet[f"C{row}"].value = music.artists
                musicSheet[f"D{row}"].value = music.duration
                musicSheet[f"E{row}"].value = music.albumImgUrl

                row += 1

        row = 2
        for offset in range(0, ExecutionLines, 100):
            response = self.database.readAllRows("Executions", limit=100, offset=offset)

            for execution in response:

                ExecutionSheet[f"A{row}"].value = execution.id
                ExecutionSheet[f"B{row}"].value = execution.musicId
                ExecutionSheet[f"C{row}"].value = execution.played_At

                row += 1

        wb.save(self.wbPath)

excel = CreateExcelFile()

print("")
excel.execute()
print("")